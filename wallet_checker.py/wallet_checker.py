import requests
from web3 import Web3
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

# Список сетей и RPC
chains = [
    { "name": "OP Mainnet", "rpc": "" },
    { "name": "Unichain", "rpc": "" },
    { "name": "Lisk", "rpc": "" },
    { "name": "Soneium", "rpc": "" },
    { "name": "Ink", "rpc": "" },
    { "name": "Base", "rpc": "" },
    { "name": "Mode", "rpc": "" },
]

# Список кошельков
wallets = [
     "0x516d5ce74c65C10908b233Fc81827E8Dbe84E40a",
     "0xe221933c6EF05B0B20106F138748282b3016ac2D",

]

def wei_to_eth(wei):
    return round(Web3.from_wei(wei, 'ether'), 6)

def get_wallet_info(rpc_url, wallet):
    try:
        web3 = Web3(Web3.HTTPProvider(rpc_url))
        if not web3.is_connected():
            return None
        
        wallet = Web3.to_checksum_address(wallet)
        balance = web3.eth.get_balance(wallet)
        tx_count = web3.eth.get_transaction_count(wallet)
        latest_block = web3.eth.block_number
        
        return {
            "balance": wei_to_eth(balance),
            "transactions": tx_count,
            "latest_block": latest_block
        }
    except Exception as e:
        print(f"Ошибка: {e}")
        return None

def save_to_excel(data):
    import math
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wallet Info"

    target_chains = [
        "OP Mainnet", "Unichain", "Lisk", "Soneium",
        "Ink", "Base", "Mode"
    ]

    chain_data = {name: [] for name in target_chains}
    for item in data:
        if item['chain'] in chain_data:
            chain_data[item['chain']].append(item)

    headers = []
    for name in target_chains:
        headers += [f"Сеть {name.split()[0]}", "Кошелёк", "Баланс", "Транзакции"]
    ws.append(headers)

    # Заливка
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")

    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Границы для заголовков
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.border = thin_border

    max_len = max(len(chain_data[name]) for name in target_chains)

    for i in range(max_len):
        row = []
        for name in target_chains:
            if i < len(chain_data[name]):
                item = chain_data[name][i]
                row += [f"№{i+1}", item["wallet"], item["balance"], item["transactions"]]
            else:
                row += ["", "", "", ""]
        ws.append(row)

        # Применяем границы к текущей строке
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin_border

 # Заливка по транзакциям и балансу
     #0–4 → 🔴 красный
     #5–19 → 🟣 фиолетовый
     #20–49 → 🔵 синий
     #50–99 → 🟠 оранжевый
     #100+ → 🟢 зелёный
    red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")       # 0-4
    purple_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")    # 5-19
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")      # 20-49
    orange_fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")    # 50-99
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")     # 100+

    for row_idx in range(2, 2 + max_len):
        for i in range(7):  # 7 сетей
            tx_cell = ws.cell(row=row_idx, column=4 + i * 4)
            if isinstance(tx_cell.value, int):
                if tx_cell.value <= 4:
                    tx_cell.fill = red_fill
                elif tx_cell.value <= 19:
                    tx_cell.fill = purple_fill
                elif tx_cell.value <= 49:
                    tx_cell.fill = blue_fill
                elif tx_cell.value <= 99:
                    tx_cell.fill = orange_fill
                else:
                    tx_cell.fill = green_fill

            balance_cell = ws.cell(row=row_idx, column=3 + i * 4)
            if balance_cell.value != "":
                balance_cell.fill = green_fill

    # Автоширина
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    filename = f"wallet_info_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    wb.save(filename)
    print(f"\n✅ Данные сохранены в файл: {filename}" )

def main():
    results = []
    for chain in chains:
        print(f"\n=== Сеть: {chain['name']} ===\n")
        for wallet in wallets:
            info = get_wallet_info(chain['rpc'], wallet)
            if info:
                print(f"Кошелек: {wallet}")
                print(f"Баланс: {info['balance']} ETH")
                print(f"Количество транзакций: {info['transactions']}")
                print("-" * 50)
                results.append({
                    "chain": chain['name'],
                    "wallet": wallet,
                    "balance": info['balance'],
                    "transactions": info['transactions'],
                    "latest_block": info['latest_block']
                })
            else:
                print(f"Не удалось подключиться к сети {chain['name']} для кошелька {wallet}")
    save_to_excel(results)

if __name__ == "__main__":
    main()

    
# библиотека openpyxl (для создания Excel-файла).
# Установи её, если ещё не установил:  pip install openpyxl
# библиотека colorama, установи её через команду:   pip install colorama
# Для запуска python wallet_checker.py
